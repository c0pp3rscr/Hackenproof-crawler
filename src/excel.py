"""Write crawled programs to a sortable/filterable ``.xlsx`` workbook."""

from __future__ import annotations

from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from crawler import Program

# (header, attribute, width, kind) — ``kind`` drives cell formatting so that
# numeric columns sort numerically and booleans render as Yes/No.
COLUMNS = [
    ("Program", "name", 34, "text"),
    ("Company", "company", 22, "text"),
    ("Reputation Required", "reputation_required", 20, "int"),
    ("KYC Required", "kyc_required", 14, "bool"),
    ("PoC Required", "poc_required", 14, "bool"),
    ("Submission Fee ($)", "submission_fee", 18, "int"),
    ("Deposit Available", "deposit_available", 18, "bool"),
    ("Max Reward", "max_reward", 16, "text"),
    ("Status", "status", 12, "text"),
    ("Activity", "activity_status", 12, "text"),
    ("Slug", "slug", 30, "text"),
    ("URL", "url", 46, "text"),
]

_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def _cell_value(program: Program, attr: str, kind: str):
    value = getattr(program, attr)
    if kind == "bool":
        if value is None:
            return ""
        return "Yes" if value else "No"
    if kind == "int":
        return value if value is not None else ""
    return value


def write_workbook(programs: list[Program], path: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "HackenProof Programs"

    # Header row.
    for col_idx, (header, _attr, _width, _kind) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")

    # Data rows.
    for row_idx, program in enumerate(programs, start=2):
        for col_idx, (_header, attr, _width, kind) in enumerate(COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx,
                    value=_cell_value(program, attr, kind))

    # Column widths.
    for col_idx, (_header, _attr, width, _kind) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    last_col = get_column_letter(len(COLUMNS))
    last_row = len(programs) + 1

    # Freeze the header and enable the filter/sort dropdowns across all columns.
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    wb.save(path)
    return path
